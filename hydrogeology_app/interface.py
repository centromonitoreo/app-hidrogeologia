import tkinter as tk
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
from hydrogeology_app.analitic_data import calculate_meq_table
from hydrogeology_app.table_management import TableManagement
from PIL import Image, ImageTk
from io import BytesIO
import base64
from tkinter import filedialog
import os

class HydrogeologyApp:
    def __init__(self, root):
        self.root = root
        self.data = None
        self.file = None
        self.treeview = None
        self.data_tree = None
        self.combobox_point = None
        self.combobox_value = None
        self.combobox_date = None
        self.combobox_bicarbonate = None
        self.combobox_calcium = None
        self.combobox_carbonate = None
        self.combobox_chlorides = None
        self.combobox_magnesium = None
        self.combobox_nitrates = None
        self.combobox_potassium = None
        self.combobox_sodium = None
        self.combobox_sulfates = None
        self.frame_parameters = None
        self.frame_columns = None
        self.frame_parameters_2 = None
        self.combobox_parameter = None
        self.canvas_frame = None
        self.frame_sheet = None
        self.combobox_sheets = None
        self.table_mannagement: TableManagement = None
        self.frame_table = None
        self.data_tree = None
        self.treeview = None
        self.df_data = pd.DataFrame()
        self.combobox_group = None
        self.combobox_color = None
        self.initialize_ui()

    def initialize_ui(self):
        with open("./data/analitica.png", "rb") as image_file:
            image_b64 = base64.b64encode(image_file.read())
        
        image = Image.open(BytesIO(base64.b64decode(image_b64)))
        self.photo = ImageTk.PhotoImage(image)
        
        self.root.wm_iconphoto(False, self.photo)
        self.root.title("HydroGeoGraph")
        self.root.geometry("1250x750")
        self.root.resizable(False, False)
        menu = tk.Menu(self.root)
        self.root.config(menu=menu)
        menu_file = tk.Menu(menu, tearoff=0)
        menu.add_cascade(label="Archivo", menu=menu_file)
        menu_file.add_command(label="Abrir..", command=self.select_file)
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=1)
        self.canvas_frame = tk.Canvas(main_frame)
        self.canvas_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        my_scrollbar = tk.Scrollbar(
            main_frame, orient=tk.VERTICAL, command=self.canvas_frame.yview
        )
        my_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas_frame.configure(yscrollcommand=my_scrollbar.set)
        self.canvas_frame.bind(
            "<Configure>",
            lambda e: self.canvas_frame.configure(
                scrollregion=self.canvas_frame.bbox("all")
            ),
        )
        frame_stace = tk.Frame(self.canvas_frame, height=30)
        self.canvas_frame.create_window((10, 30), window=frame_stace, anchor="nw")
        self.frame_sheet = tk.Frame(self.canvas_frame, height=30, pady=10, padx=10)
        self.combobox_sheets = self.generate_combobox(
            self.frame_sheet, "Seleccionar Pestaña: ", 0, 0
        )
        self.create_button(self.frame_sheet, "Leer Pestaña", self.select_sheet, 2, 0)
        self.canvas_frame.create_window((10, 20), window=self.frame_sheet, anchor="nw")
        self.ajustar_ypadx(self.frame_sheet, 4)
        frame_titulo_parametros = tk.Frame(self.canvas_frame, height=30, padx=10)
        label_titulo_parametro = tk.Label(
            frame_titulo_parametros, text="Seleccionar Configureción de Columnas: "
        )
        label_titulo_parametro.grid(row=0, column=0, sticky="w")
        self.canvas_frame.create_window(
            (10, 140), window=frame_titulo_parametros, anchor="nw"
        )
        self.frame_columns = tk.Frame(self.canvas_frame, height=30, padx=10)
        self.frame_columns.grid(row=2, column=0, sticky="w")
        self.combobox_point = self.generate_combobox(
            self.frame_columns, "Columna puntos: ", 1, 0, "Horizontal"
        )
        self.combobox_parameter = self.generate_combobox(
            self.frame_columns, "Columna parametros: ", 2, 0, "Horizontal"
        )
        self.combobox_value = self.generate_combobox(
            self.frame_columns, "Columna Valor: ", 3, 0, "Horizontal"
        )
        self.combobox_date = self.generate_combobox(
            self.frame_columns, "Columna Fecha: ", 4, 0, "Horizontal"
        )
        boton_param = tk.Button(
            self.frame_columns,
            text="Leer Columnas",
            command=self.select_parameter_labels,
        )
        boton_param.grid(row=5, column=0, sticky="w")
        self.canvas_frame.create_window(
            (10, 170), window=self.frame_columns, anchor="nw"
        )
        self.ajustar_ypadx(self.frame_columns, 4)
        frame_titulo_parametros = tk.Frame(self.canvas_frame, height=30, padx=10)
        label_titulo_parametro = tk.Label(
            frame_titulo_parametros, text="Seleccionar Configureción de Parametros: "
        )
        label_titulo_parametro.grid(row=0, column=0, sticky="w")
        self.canvas_frame.create_window(
            (400, 140), window=frame_titulo_parametros, anchor="nw"
        )
        self.frame_parameters = tk.Frame(self.canvas_frame, height=30, padx=10)
        self.combobox_bicarbonate = self.generate_combobox(
            self.frame_parameters, "Etiqueta Bicarbonato (mg/L): ", 1, 0, "Horizontal"
        )
        self.combobox_calcium = self.generate_combobox(
            self.frame_parameters, "Etiqueta Calcio (mg/L): ", 2, 0, "Horizontal"
        )
        self.combobox_carbonate = self.generate_combobox(
            self.frame_parameters, "Etiqueta Carbonato (mg/L): ", 3, 0, "Horizontal"
        )
        self.combobox_chlorides = self.generate_combobox(
            self.frame_parameters, "Etiqueta Cloruros (mg/L Cl-): ", 4, 0, "Horizontal"
        )
        self.combobox_magnesium = self.generate_combobox(
            self.frame_parameters, "Etiqueta Magnesio (mg/L): ", 5, 0, "Horizontal"
        )
        boton_table = tk.Button(
            self.frame_parameters,
            text="Leer Columnas",
            command=self.generate_table,
        )
        boton_table.grid(row=10, column=0, sticky="w")
        self.canvas_frame.create_window(
            (400, 170), window=self.frame_parameters, anchor="nw"
        )

        self.frame_parameters_2 = tk.Frame(self.canvas_frame, height=30, padx=10)
        self.combobox_nitrates = self.generate_combobox(
            self.frame_parameters_2,
            "Etiqueta Nitratos (mg/L N-NO3): ",
            1,
            0,
            "Horizontal",
        )
        self.combobox_potassium = self.generate_combobox(
            self.frame_parameters_2, "Etiqueta Potasio (mg/L): ", 2, 0, "Horizontal"
        )
        self.combobox_sodium = self.generate_combobox(
            self.frame_parameters_2, "Etiqueta Sodio (mg/L): ", 3, 0, "Horizontal"
        )
        self.combobox_sulfates = self.generate_combobox(
            self.frame_parameters_2,
            "Etiqueta Sulfatos (mg/L SO4-2): ",
            4,
            0,
            "Horizontal",
        )
        self.canvas_frame.create_window(
            (800, 170), window=self.frame_parameters_2, anchor="nw"
        )
        self.ajustar_ypadx(self.frame_parameters, 3)
        self.ajustar_ypadx(self.frame_parameters_2, 3)
        self.table_mannagement = TableManagement(self, pd.DataFrame())
        self.table_mannagement.generate_table()
        self.root.mainloop()

    def populate_combo_frame(self, frame, values):
        for widget in frame.winfo_children():
            if isinstance(widget, ttk.Combobox):
                widget["values"] = []
                widget["values"] = values

    def set_value_combo(self, combobox, parameter):
        if parameter in combobox["values"]:
            combobox.set(parameter)

    def clean_frame(self, frames):
        for frame in frames:
            for widget in frame.winfo_children():
                if isinstance(widget, ttk.Combobox):
                    widget["values"] = []
                    widget.set("")
                elif isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)

    def generate_combobox(self, frame, text_label, row, col, positional_label=None):
        label = tk.Label(frame, text=text_label)
        label.grid(row=row, column=col, sticky="w")
        combobox = ttk.Combobox(frame)
        if positional_label == "Horizontal":
            combobox.grid(row=row, column=col + 1, sticky="w")
        else:
            combobox.grid(row=row + 1, column=col, sticky="w")
        combobox.config(width=30)
        return combobox

    def create_button(self, frame, text, command, row, col):
        boton = tk.Button(frame, text=text, command=command)
        boton.grid(row=row, column=col, sticky="w")

    def select_file(self):
        self.clean_frame(
            [self.frame_sheet, self.frame_columns, self.frame_parameters]
        )
        self.file = filedialog.askopenfilename(
            filetypes=[("Archivos de Excel", "*.xlsx")]
        )
        
        if not os.path.exists(self.file):
            return
        workbook = load_workbook(filename=self.file)
        sheets_names = workbook.sheetnames
        self.populate_combo_frame(self.frame_sheet, sheets_names)

    def select_sheet(self):
        self.clean_frame([self.frame_columns, self.frame_parameters])
        sheet_name = self.combobox_sheets.get()
        self.data = pd.read_excel(self.file, sheet_name=sheet_name)
        columns = self.data.columns.tolist()
        self.populate_combo_frame(self.frame_columns, columns)

    def select_parameter_labels(self):
        self.clean_frame([self.frame_parameters, self.frame_parameters_2])
        filled_columns = self.check_completion_frame(self.frame_columns, "Columnas")
        if filled_columns & self.check_date_columns() & self.check_value_column():
            parameters = (
                self.data[[self.combobox_parameter.get()]]
                .sort_values(by=self.combobox_parameter.get())[
                    self.combobox_parameter.get()
                ]
                .unique()
                .tolist()
            )
            self.populate_combo_frame(self.frame_parameters, parameters)
            self.populate_combo_frame(self.frame_parameters_2, parameters)
            self.set_value_combo(self.combobox_bicarbonate, "Bicarbonato (mg/L)")
            self.set_value_combo(self.combobox_calcium, "Calcio (mg/L)")
            self.set_value_combo(self.combobox_carbonate, "Carbonato (mg/L)")
            self.set_value_combo(self.combobox_chlorides, "Cloruros (mg/L Cl-)")
            self.set_value_combo(self.combobox_magnesium, "Magnesio (mg/L)")
            self.set_value_combo(self.combobox_nitrates, "Nitratos (mg/L N-NO3)")
            self.set_value_combo(self.combobox_potassium, "Potasio (mg/L)")
            self.set_value_combo(self.combobox_sodium, "Sodio (mg/L)")
            self.set_value_combo(self.combobox_sulfates, "Sulfatos (mg/L SO4-2)")
           

    def check_completion_frame(self, frame, name_frame):
        for widget in frame.winfo_children():
            if isinstance(widget, ttk.Combobox) | isinstance(widget, tk.Entry):
                if len(widget.get()) == 0:
                    tk.messagebox.showinfo(
                        "Mensaje de Alerta",
                        f"Diligenciar todos los campos del frame {name_frame}",
                    )
                    return False
            return True

    def check_date_columns(self):
        data_copy = self.data.copy()
        col_date = self.combobox_date.get()
        date_format = "%d/%m/%Y"
        try:
            data_copy[col_date] = pd.to_datetime(
                data_copy[col_date], format=date_format
            )
            print(data_copy[col_date])
            return True
        except KeyError:
            tk.messagebox.showinfo(
                "Mensaje de Alerta", f"La columna {col_date} no existe en los datos."
            )
            return False
        except ValueError:
            tk.messagebox.showinfo(
                "Mensaje de Alerta",
                f"La columna {col_date} no cuenta con el formato {date_format}.",
            )
            return False

    def check_value_column(self):
        data_copy = self.data.copy()
        col_valor = self.combobox_value.get()
        try:
            data_copy[col_valor].astype(float)
            return True
        except KeyError:
            tk.messagebox.showinfo(
                "Mensaje de Alerta", f"La columna {col_valor} no existe en los datos."
            )
            return False
        except ValueError:
            tk.messagebox.showinfo(
                "Mensaje de Alerta",
                f"La columna {col_valor} no cuenta con el formato numérico.",
            )
            return False

    def generate_table(self):
        self.check_completion_frame(self.frame_parameters, "Parametros")
        self.check_completion_frame(self.frame_columns, "Columnas")
        para_sulfates = (
            self.combobox_sulfates.get()
            if len(self.combobox_sulfates.get()) > 0
            else "null_sulfatos"
        )
        para_sodium = (
            self.combobox_sodium.get()
            if len(self.combobox_sodium.get()) > 0
            else "null_sodio"
        )
        para_potassium = (
            self.combobox_potassium.get()
            if len(self.combobox_potassium.get()) > 0
            else "null_potasio"
        )
        para_nitrates = (
            self.combobox_nitrates.get()
            if len(self.combobox_nitrates.get()) > 0
            else "null_nitratos"
        )
        para_magnesium = (
            self.combobox_magnesium.get()
            if len(self.combobox_magnesium.get()) > 0
            else "null_magnesio"
        )
        para_chlorides = (
            self.combobox_chlorides.get()
            if len(self.combobox_chlorides.get()) > 0
            else "null_cloruros"
        )
        para_carbonate = (
            self.combobox_carbonate.get()
            if len(self.combobox_carbonate.get()) > 0
            else "null_carbonato"
        )
        para_calcium = (
            self.combobox_calcium.get()
            if len(self.combobox_calcium.get()) > 0
            else "null_calcio"
        )
        para_bicarbonate = (
            self.combobox_bicarbonate.get()
            if len(self.combobox_bicarbonate.get()) > 0
            else "null_bicarbonato"
        )
        keys_repetidos = self.find_duplicates(
            [
                para_sulfates,
                para_sodium,
                para_potassium,
                para_nitrates,
                para_magnesium,
                para_chlorides,
                para_carbonate,
                para_calcium,
                para_bicarbonate,
            ]
        )
        if len(keys_repetidos) > 0:
            tk.messagebox.showerror(
                "Error en parametros",
                f"Selecciono dos parametros con el mismo nombre {keys_repetidos}",
            )
            return None
        dict_rename = {
            para_sulfates: "Sulfatos (mg/L)",
            para_sodium: "Sodio (mg/L)",
            para_potassium: "Potasio (mg/L)",
            para_nitrates: "Nitratos (mg/L)",
            para_magnesium: "Magnesio (mg/L)",
            para_chlorides: "Cloruros (mg/L)",
            para_carbonate: "Carbonato (mg/L)",
            para_calcium: "Calcio (mg/L)",
            para_bicarbonate: "Bicarbonato (mg/L)",
        }
        self.table_mannagement.df_data = calculate_meq_table(
            self.data,
            dict_rename,
            self.combobox_parameter.get(),
            self.combobox_value.get(),
        )
        self.table_mannagement.generate_table()

    def find_duplicates(self, lst):
        unique_elements = set()
        duplicates = set()
        for element in lst:
            if element in unique_elements:
                duplicates.add(element)
            else:
                unique_elements.add(element)
        return list(duplicates)

    def ajustar_ypadx(self, frame, nuevo_ypadx):
        for widget in frame.winfo_children():
            widget.grid_configure(pady=nuevo_ypadx)

    def ajustar_xpadx(self, frame, nuevo_xpadx):
        for widget in frame.winfo_children():
            widget.grid_configure(padx=nuevo_xpadx)

